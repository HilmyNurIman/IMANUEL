# data_handler.py - milik Hanif
# Semua operasi terkait file Excel dan data peminjaman

from openpyxl import Workbook, load_workbook
from tabulate import tabulate
from datetime import datetime
import os

DATA_FILE = "peminjaman_laptop.xlsx"

def init_file_excel():
    """Buat file Excel baru jika belum ada"""
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Peminjaman Laptop"
        ws.append(["Nama", "Kelas", "Tanggal Pinjam", "Tanggal Kembali", "Keterangan"])
        wb.save(DATA_FILE)

# Jalankan sekali saat modul diimport
init_file_excel()

def baca_data_excel():
    """Baca semua data dari Excel (mulai baris 2)"""
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            data.append({
                "Nama": row[0],
                "Kelas": row[1],
                "Tanggal Pinjam": row[2],
                "Tanggal Kembali": row[3],
                "Keterangan": row[4]
            })
    return data

def tambah_data_excel(nama, kelas, tgl_pinjam, tgl_kembali=None, keterangan=""):
    """Tambah satu baris data baru ke Excel"""
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        ws.append([nama, kelas, tgl_pinjam, tgl_kembali, keterangan])
        wb.save(DATA_FILE)
        return True
    except Exception as e:
        print(f"Error saat menambah data: {e}")
        return False

def hapus_semua_data_excel():
    """Hapus semua data (kecuali header)"""
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
        wb.save(DATA_FILE)

def tampilkan_tabel(data):
    """Tampilkan data dalam format tabel cantik"""
    if not data:
        print("📂 Data masih kosong\n")
        return

    table = [[
        d["Nama"], d["Kelas"], d["Tanggal Pinjam"],
        d["Tanggal Kembali"] or "-", d["Keterangan"]
    ] for d in data]

    print("\n📊 DATA PEMINJAMAN LAPTOP\n")
    print(tabulate(
        table,
        headers=["Nama", "Kelas", "Pinjam", "Kembali", "Keterangan"],
        tablefmt="grid",
    ))
    print()

def update_status_kembali(nama, kelas):
    """Update kolom Tanggal Kembali untuk peminjam tertentu"""
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    
    nama_norm = nama.strip().title()
    kelas_norm = kelas.strip().upper()
    
    kandidat = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        nama_excel = str(row[0].value or "").strip().title()
        kelas_excel = str(row[1].value or "").strip().upper()
        
        if nama_excel == nama_norm and kelas_excel.startswith(kelas_norm):
            kandidat.append({
                "row": row_idx,
                "nama": nama_excel,
                "kelas": kelas_excel,
                "pinjam": row[2].value,
                "keterangan": row[4].value or "-"
            })
    
    if not kandidat:
        return False, "Tidak ditemukan data yang cocok"
    
    if len(kandidat) == 1:
        row_idx = kandidat[0]["row"]
        ws.cell(row=row_idx, column=4).value = datetime.now().strftime("%Y-%m-%d %H:%M")
        wb.save(DATA_FILE)
        return True, f"Diupdate baris {row_idx}"
    
    # Jika banyak kandidat
    print("\nDitemukan beberapa entri serupa:")
    for i, item in enumerate(kandidat, 1):
        print(f"{i}. Baris {item['row']}: {item['nama']} | {item['kelas']} | Pinjam: {item['pinjam']} | Ket: {item['keterangan']}")
    
    pilihan = input("\nNomor mana yang mau diupdate? (angka atau 'batal'): ").strip()
    if pilihan.lower() == 'batal':
        return False, "Dibatalkan oleh user"
    
    try:
        idx = int(pilihan) - 1
        if 0 <= idx < len(kandidat):
            row_idx = kandidat[idx]["row"]
            ws.cell(row=row_idx, column=4).value = datetime.now().strftime("%Y-%m-%d %H:%M")
            wb.save(DATA_FILE)
            return True, f"Berhasil update baris {row_idx}"
        else:
            return False, "Nomor tidak valid"
    except ValueError:
        return False, "Input bukan angka valid"